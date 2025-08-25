#!/usr/bin/env python3

import argparse
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

def parse_hash_file(file_path):
    """
    Parse the hash file and extract relevant information
    """
    data = []

    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
            for line in file:
                line = line.strip()
                if not line:
                    continue

                # Split by colon
                parts = line.split(':')
                if len(parts) >= 4:
                    # Extract domain\username
                    domain_user = parts[0]
                    if '\\' in domain_user:
                        domain, username = domain_user.split('\\', 1)
                    else:
                        domain = ""
                        username = domain_user

                    # Extract other fields
                    uid = parts[1] if len(parts) > 1 else ""
                    lm_hash = parts[2] if len(parts) > 2 else ""
                    nt_hash = parts[3] if len(parts) > 3 else ""

                    data.append({
                        'Domain': domain,
                        'Username': username,
                        'UID': uid,
                        'LM Hash': lm_hash,
                        'NT Hash': nt_hash,
                        'Full Entry': line
                    })
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
        return []
    except Exception as e:
        print(f"Error reading hash file: {e}")
        return []

    return data

def parse_cracked_file(file_path):
    """
    Parse the cracked passwords file and return a dictionary of hash:password
    """
    cracked_passwords = {}

    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
            for line in file:
                line = line.strip()
                if not line:
                    continue

                # Skip lines with [not found]
                if '[not found]' in line.lower():
                    continue

                # Handle different formats
                password = None
                hash_value = None

                # Format: hash:password
                if ':' in line:
                    parts = line.split(':', 1)
                    if len(parts) == 2:
                        hash_value = parts[0].strip().lower()
                        password = parts[1].strip()
                # Format: hash password (space separated)
                elif ' ' in line:
                    parts = line.split(' ', 1)
                    if len(parts) == 2:
                        hash_value = parts[0].strip().lower()
                        password = parts[1].strip()

                # Only add if both hash and password are valid
                if hash_value and password and '[not found]' not in password.lower():
                    cracked_passwords[hash_value] = password
    except FileNotFoundError:
        print(f"Error: Cracked passwords file '{file_path}' not found.")
        return {}
    except Exception as e:
        print(f"Error reading cracked file: {e}")
        return {}

    return cracked_passwords

def match_passwords(hash_data, cracked_passwords):
    """
    Match cracked passwords with hash data
    """
    matched_data = []

    for entry in hash_data:
        nt_hash = entry['NT Hash'].lower()
        lm_hash = entry['LM Hash'].lower()

        # Check if NT hash or LM hash exists in cracked passwords
        password = None
        if nt_hash in cracked_passwords:
            password = cracked_passwords[nt_hash]
        elif lm_hash in cracked_passwords:
            password = cracked_passwords[lm_hash]

        # Only include entries where password was found
        if password:
            matched_data.append({
                'Domain': entry['Domain'],
                'Username': entry['Username'],
                'Password': password
            })

    return matched_data

def apply_styling_to_sheet(worksheet, has_data=True):
    """
    Apply Titillium Web font styling and formatting to Excel sheet
    """
    if not has_data:
        return
        
    # Apply font to header row
    header_font = Font(name='Titillium Web', size=11, bold=True)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    # Style headers
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Apply font to all data cells
    data_font = Font(name='Titillium Web', size=10)
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = data_font
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def save_to_excel(hash_data, matched_data, output_file):
    """
    Save to Excel file with exactly 2 sheets and Titillium Web font styling
    """
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Sheet 1: All parsed hashes
            if hash_data:
                df_hashes = pd.DataFrame(hash_data)
                df_hashes.to_excel(writer, sheet_name='All_Hashes', index=False)
                print(f"Sheet 1 - All_Hashes: {len(hash_data)} entries")
            else:
                # Create empty sheet if no data
                pd.DataFrame([['No hash data found']], columns=['Message']).to_excel(writer, sheet_name='All_Hashes', index=False)
                print("Sheet 1 - All_Hashes: 0 entries")

            # Sheet 2: Only matched passwords
            if matched_data:
                df_matched = pd.DataFrame(matched_data)
                df_matched.to_excel(writer, sheet_name='Cracked_Passwords', index=False)
                print(f"Sheet 2 - Cracked_Passwords: {len(matched_data)} entries")
            else:
                # Create empty sheet if no matches
                pd.DataFrame([['No cracked passwords found']], columns=['Message']).to_excel(writer, sheet_name='Cracked_Passwords', index=False)
                print("Sheet 2 - Cracked_Passwords: 0 entries")

        # Re-open and apply styling
        workbook = load_workbook(output_file)
        
        # Style All_Hashes sheet
        if hash_data:
            apply_styling_to_sheet(workbook['All_Hashes'], has_data=True)
        else:
            apply_styling_to_sheet(workbook['All_Hashes'], has_data=False)
            
        # Style Cracked_Passwords sheet
        if matched_data:
            apply_styling_to_sheet(workbook['Cracked_Passwords'], has_data=True)
        else:
            apply_styling_to_sheet(workbook['Cracked_Passwords'], has_data=False)
        
        # Save styled workbook
        workbook.save(output_file)
        print(f"Excel file successfully created with Titillium Web font: {output_file}")
        
    except Exception as e:
        print(f"Error saving to Excel: {e}")

def main():
    parser = argparse.ArgumentParser(description='Parse hash file and match with cracked passwords (Styled with Titillium Web)')
    parser.add_argument('-f', '--file', required=True, help='Input hash file')
    parser.add_argument('-p', '--passwords', required=True, help='Cracked passwords file')
    parser.add_argument('-o', '--output', required=True, help='Output Excel file')

    args = parser.parse_args()

    # Parse the hash file
    print(f"Parsing hash file: {args.file}")
    hash_data = parse_hash_file(args.file)
    print(f"Found {len(hash_data)} hash entries")

    # Parse the cracked passwords file
    print(f"Parsing cracked passwords file: {args.passwords}")
    cracked_passwords = parse_cracked_file(args.passwords)
    print(f"Found {len(cracked_passwords)} cracked passwords (excluding [not found])")

    # Match passwords with hashes
    print("Matching passwords with hashes...")
    matched_data = match_passwords(hash_data, cracked_passwords)
    print(f"Successfully matched {len(matched_data)} passwords")

    # Save to Excel with 2 sheets and styling
    save_to_excel(hash_data, matched_data, args.output)

if __name__ == "__main__":
    main()
